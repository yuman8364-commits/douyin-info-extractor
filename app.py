# -*- coding: utf-8 -*-
"""抖音信息提取工具 - Windows 桌面界面（Tkinter）。"""

import json
import logging
from logging.handlers import RotatingFileHandler
import os
import queue
import random
import re
import shutil
import subprocess
import sys
import threading
import time
import traceback
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk

import exporter
import extractor
import input_parser
from storage import ArtifactTransaction, RecordDeletionTransaction, TransactionError
from tasking import TaskCancelled, TaskMessage, ensure_not_cancelled, interruptible_wait
from openpyxl import load_workbook
from PIL import Image, ImageTk

APP_VERSION = "2.0.21"
PREVIEW_BOX_SIZE = (190, 250)
PREVIEW_IMAGE_SIZE = (170, 230)
PREVIEW_BACKGROUND = (242, 242, 242, 255)
SOURCE_DIR = Path(__file__).resolve().parent
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else SOURCE_DIR


def _writable_state_dir() -> Path:
    """优先使用便携 data；只读时回退到当前用户的 LocalAppData。"""
    preferred = APP_DIR / "data" if getattr(sys, "frozen", False) else SOURCE_DIR
    candidates = [preferred]
    if getattr(sys, "frozen", False):
        local_app_data = os.environ.get("LOCALAPPDATA")
        if local_app_data:
            candidates.append(Path(local_app_data) / "抖音信息提取工具")
    last_error: OSError | None = None
    for candidate in candidates:
        probe = candidate / ".write_test"
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return candidate
        except OSError as exc:
            last_error = exc
            try:
                probe.unlink(missing_ok=True)
            except OSError:
                pass
    raise OSError(f"程序状态目录不可写：{last_error}")


STATE_DIR = _writable_state_dir()
CONFIG_FILE = STATE_DIR / "config.json"
INPUT_CACHE_FILE = STATE_DIR / "input_cache.txt"
BROWSER_PROFILE_DIR = STATE_DIR / "browser_profile"
STARTUP_ERROR_FILE = STATE_DIR / "启动错误.log"
LOG_NAME = "提取日志.log"
DIVIDER_RE = input_parser.DIVIDER_RE
SPREADSHEET_PROCESS_NAMES = ("EXCEL.EXE", "et.exe")


# 输入解析逻辑已拆到纯逻辑模块；保留这些公开名称以兼容旧测试和调用。
split_entry_blocks = input_parser.split_entry_blocks
_block_raw_content = input_parser._block_raw_content
_is_placeholder_block = input_parser._is_placeholder_block
build_input_jobs = input_parser.build_input_jobs
input_job_at_line = input_parser.input_job_at_line
normalize_input_text = input_parser.normalize_input_text


def prepare_preview_image(path) -> Image.Image:
    """把封面等比例居中到固定画布，避免 Tk Label 切换图片后挤压布局。"""
    with Image.open(path) as source:
        image = source.copy()
    if image.mode not in ("RGB", "RGBA"):
        image = image.convert("RGBA")
    image.thumbnail(PREVIEW_IMAGE_SIZE, Image.Resampling.LANCZOS)
    canvas = Image.new("RGBA", PREVIEW_IMAGE_SIZE, PREVIEW_BACKGROUND)
    offset = (
        (PREVIEW_IMAGE_SIZE[0] - image.width) // 2,
        (PREVIEW_IMAGE_SIZE[1] - image.height) // 2,
    )
    if image.mode == "RGBA":
        canvas.paste(image, offset, image)
    else:
        canvas.paste(image, offset)
    return canvas.convert("RGB")


def build_cover_map(output_dir: Path, seqs) -> dict[int, str]:
    """按「封面/N.*」为给定序号建立封面路径映射。"""
    cover_map: dict[int, str] = {}
    cover_dir = output_dir / "封面"
    if cover_dir.exists():
        for seq in seqs:
            matches = sorted(cover_dir.glob(f"{seq}.*"))
            if matches:
                cover_map[seq] = str(matches[0])
    return cover_map


def load_config() -> dict:
    try:
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8").lstrip("\ufeff"))
    except Exception:
        return {}


def load_input_cache() -> str:
    """读取上次关闭工具时保存的输入框内容（之前粘贴的链接）。"""
    try:
        return INPUT_CACHE_FILE.read_text(encoding="utf-8").lstrip("\ufeff").rstrip("\n")
    except Exception:
        return ""


def save_config(updates: dict) -> None:
    """合并写入配置，避免覆盖已有的其它设置。"""
    try:
        config = load_config()
        config.update(updates)
        CONFIG_FILE.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


def force_close_spreadsheet_apps(logger: logging.Logger | None = None) -> list[str]:
    """强制关闭全部 Microsoft Excel/WPS 表格进程。"""
    closed: list[str] = []
    creation_flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    for image_name in SPREADSHEET_PROCESS_NAMES:
        try:
            result = subprocess.run(
                ["taskkill", "/F", "/IM", image_name],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
                creationflags=creation_flags,
            )
        except OSError as exc:
            if logger is not None:
                logger.warning("强制关闭 %s 失败：%s", image_name, exc)
            continue
        if result.returncode == 0:
            closed.append(image_name)
    return closed


def update_records_force_close(
    path,
    records: dict[int, dict],
    seqs: list[int] | None = None,
    cover_map: dict[int, str] | None = None,
    *,
    allow_force_close: bool = True,
    keep_backup: bool = False,
) -> None:
    """工作簿被占用时按用户要求强制关闭 Excel/WPS 并重试一次。"""
    try:
        exporter.update_records(
            path, records, seqs, cover_map, keep_backup=keep_backup
        )
        return
    except exporter.WorkbookInUseError:
        if not allow_force_close:
            raise

    logger = logging.getLogger("douyin_tool")
    logger.warning(
        "提取记录.xlsx 被占用，按用户设置强制关闭全部 Excel/WPS 表格进程后重试写入"
    )
    closed = force_close_spreadsheet_apps(logger)
    logger.warning("已结束的表格进程：%s", "、".join(closed) if closed else "未检测到或已退出")
    time.sleep(0.8)
    exporter.update_records(path, records, seqs, cover_map, keep_backup=keep_backup)


def delete_record_force_close(
    path, seq: int, *, keep_backup: bool = False
) -> bool:
    """删除行遇占用时强制关闭 Excel/WPS 并重试一次。"""
    try:
        return exporter.delete_record(path, seq, keep_backup=keep_backup)
    except exporter.WorkbookInUseError:
        pass

    logger = logging.getLogger("douyin_tool")
    logger.warning(
        "删除记录时提取记录.xlsx 被占用，强制关闭全部 Excel/WPS 后重试"
    )
    closed = force_close_spreadsheet_apps(logger)
    logger.warning("已结束的表格进程：%s", "、".join(closed) if closed else "未检测到或已退出")
    time.sleep(0.8)
    return exporter.delete_record(path, seq, keep_backup=keep_backup)


def fetch_with_retry(
    logger,
    seq: int,
    link: str,
    cancel_event: threading.Event | None = None,
    access_context: extractor.AccessContext | None = None,
):
    """统一有限重试；浏览器验证失败向上抛出以触发批次熔断。"""
    last_error = "获取失败"
    for attempt in (1, 2):
        ensure_not_cancelled(cancel_event)
        try:
            if access_context is not None:
                return access_context.fetch_record(link), None
            return extractor.fetch_record(link, cancel_event), None
        except TaskCancelled:
            raise
        except extractor.BrowserVerificationError:
            raise
        except extractor.InvalidLinkError:
            return None, "链接无效"
        except extractor.TargetUnavailableError:
            return None, "目标作品已失效（浏览器自动跳转到其他作品）"
        except extractor.LoginRequiredError as exc:
            last_error = f"目标作品暂不可用（{exc}）"
        except extractor.PageStructureError as exc:
            last_error = f"获取失败（页面结构可能已变化：{exc}）"
        except extractor.CaptchaChallengeError as exc:
            last_error = f"风控或验证异常（{exc}）"
        except extractor.WafBlockedError as exc:
            last_error = f"风控或网络异常（{exc}）"
        except Exception as exc:
            last_error = f"获取失败（{exc}）"
        if attempt == 1:
            logger.warning("顺序 %d：%s，有限重试一次", seq, last_error)
            interruptible_wait(2 + random.random() * 2, cancel_event)
    return None, last_error


build_input_tasks = input_parser.build_input_tasks


def default_output_dir() -> str:
    desktop = Path.home() / "Desktop"
    return str(desktop if desktop.exists() else APP_DIR)


def next_sequence(output_dir: str | Path) -> int:
    """综合媒体、文案、封面和工作簿返回下一个可用编号。"""
    root = Path(output_dir)
    numbers: set[int] = set()

    videos = root / "爆款视频"
    if videos.exists():
        for entry in videos.iterdir():
            match = re.match(r"^(\d+)(?:\.mp4)?$", entry.name)
            if match:
                numbers.add(int(match.group(1)))

    captions = root / "文案提取"
    if captions.exists():
        for entry in captions.iterdir():
            match = re.match(r"^(\d+)\.txt$", entry.name)
            if match:
                numbers.add(int(match.group(1)))

    covers = root / "封面"
    if covers.exists():
        for entry in covers.iterdir():
            match = re.match(r"^(\d+)\.[A-Za-z0-9]+$", entry.name)
            if match:
                numbers.add(int(match.group(1)))

    try:
        numbers.update(exporter.read_records(root / "提取记录.xlsx"))
    except Exception:
        # 工作簿异常由正式任务流程报告；编号计算仍避免因此崩溃。
        pass

    return (max(numbers) + 1) if numbers else 1


def setup_logger(log_path: Path) -> logging.Logger:
    """建立带轮转的共享日志，并关闭上一批次留下的文件句柄。"""
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    for handler in list(root.handlers):
        root.removeHandler(handler)
        try:
            handler.close()
        except Exception:
            pass
    handler = RotatingFileHandler(
        log_path, maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8"
    )
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    root.addHandler(handler)
    root.info("抖音信息提取工具 %s", APP_VERSION)
    return root


def find_same_size_file(
    folder: Path, size: int, exclude: Path | None = None, exclude_dir: Path | None = None
) -> str | None:
    """在 folder 下递归查找字节数等于 size 的文件，返回相对路径；没有则返回 None。

    exclude：排除单个文件；exclude_dir：排除整个子目录（本次新下载的文件不参与比较）。
    用于重复提取检测：文件大小完全相同，大概率是同一作品的重复提取。
    """
    if not folder.exists() or size <= 0:
        return None
    excluded = exclude.resolve() if exclude is not None else None
    excluded_dir = exclude_dir.resolve() if exclude_dir is not None else None
    for entry in folder.rglob("*"):
        if not entry.is_file():
            continue
        if excluded_dir is not None and entry.is_relative_to(excluded_dir):
            continue
        if excluded is not None and entry.resolve() == excluded:
            continue
        try:
            if entry.stat().st_size == size:
                return str(entry.relative_to(folder))
        except OSError:
            continue
    return None


def scan_seq_numbers(output_dir: str | Path) -> tuple[set[int], set[int], set[int]]:
    """扫描输出目录，返回 (视频编号集合, 文案编号集合, 封面编号集合)。"""
    root = Path(output_dir)
    videos: set[int] = set()
    captions: set[int] = set()
    covers: set[int] = set()

    vdir = root / "爆款视频"
    if vdir.exists():
        for entry in vdir.iterdir():
            match = re.match(r"^(\d+)(?:\.mp4)?$", entry.name)
            if match:
                videos.add(int(match.group(1)))

    cdir = root / "文案提取"
    if cdir.exists():
        for entry in cdir.iterdir():
            match = re.match(r"^(\d+)\.txt$", entry.name)
            if match:
                captions.add(int(match.group(1)))

    cover_dir = root / "封面"
    if cover_dir.exists():
        for entry in cover_dir.iterdir():
            match = re.match(r"^(\d+)\.[A-Za-z0-9]+$", entry.name)
            if match:
                covers.add(int(match.group(1)))
    return videos, captions, covers


def scan_note_numbers(output_dir: str | Path) -> set[int]:
    """扫描输出目录「爆款视频」下以序号命名的目录（图文图集），返回序号集合。"""
    root = Path(output_dir)
    vdir = root / "爆款视频"
    notes: set[int] = set()
    if vdir.exists():
        for entry in vdir.iterdir():
            if entry.is_dir() and re.fullmatch(r"\d+", entry.name):
                notes.add(int(entry.name))
    return notes


def _check_excel_order(output_dir: Path, videos: set[int]) -> bool:
    """检查表格「顺序」列是否与视频文件编号一致；不一致返回 True。"""
    path = output_dir / "提取记录.xlsx"
    if not path.exists() or not videos:
        return False
    try:
        workbook = load_workbook(path, read_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if "顺序" not in headers:
            workbook.close()
            return False
        seq_index = headers.index("顺序")
        seqs: set[int] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            values = list(row)
            value = values[seq_index] if seq_index < len(values) else None
            if isinstance(value, int) and value > 0:
                seqs.add(value)
        workbook.close()
    except Exception:
        return False
    return bool(seqs) and seqs != videos


class DouyinExtractorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.message_queue: queue.Queue[TaskMessage] = queue.Queue()
        self.worker: threading.Thread | None = None
        self.refresh_thread: threading.Thread | None = None
        self.cancel_event = threading.Event()
        self.close_requested = False
        self._retry_request = False
        self.retry_run = False
        self.running = False
        self.single_run = False
        self.refreshing = False
        self.auto_refresh = False
        self.records: dict[str, dict] = {}
        self.ok_count = 0
        self.fail_count = 0
        self.refresh_count = 0
        self.cancel_count = 0
        self.rollback_count = 0
        self.unchecked_count = 0
        self.pause_message = ""
        self.failed_jobs: list[tuple[int | None, str]] = []
        self.thumb_ref = None
        self.output_dir = Path(load_config().get("output_dir") or default_output_dir())
        self._build_ui()
        # 恢复上次关闭前粘贴的链接，光标停在最新编号处，可直接继续粘贴
        cached = load_input_cache()
        if cached.strip() and cached.strip() != "1.":
            self.input_text.delete("1.0", "end")
            self.input_text.insert("1.0", cached)
        self._enforce_input_sequences()
        self.input_text.mark_set("insert", "end-1c")
        self.input_text.see("insert")
        self.load_existing_records()

    def _post(self, kind: str, payload=None, extra=None) -> None:
        self.message_queue.put(TaskMessage(kind, payload, extra))

    def _save_input_cache(self) -> None:
        """把输入框当前内容（之前粘贴的链接）保存到缓存文件。"""
        try:
            content = self.input_text.get("1.0", "end").rstrip("\n")
            INPUT_CACHE_FILE.write_text(content + "\n", encoding="utf-8")
        except Exception:
            pass

    def _on_close(self) -> None:
        self._save_input_cache()
        active = self.running or self.refreshing or any(
            thread is not None and thread.is_alive()
            for thread in (self.worker, self.refresh_thread)
        )
        if active:
            if not self.close_requested:
                if not messagebox.askyesno(
                    "停止任务并退出",
                    "当前任务仍在运行。是否安全停止任务并在清理完成后退出？",
                    parent=self.root,
                ):
                    return
                self.close_requested = True
                self.stop_task()
            self.status_var.set("正在停止任务并清理临时文件，请稍候…")
            self.root.after(100, self._wait_then_close)
            return
        self.root.destroy()

    def _wait_then_close(self) -> None:
        threads = (self.worker, self.refresh_thread)
        if any(thread is not None and thread.is_alive() for thread in threads):
            self.root.after(100, self._wait_then_close)
            return
        self.root.destroy()

    def _set_task_controls(self, enabled: bool) -> None:
        """统一启用/禁用主操作及会改变任务输入或输出目录的控件。"""
        state = "normal" if enabled else "disabled"
        for button in (
            self.start_button,
            self.single_button,
            self.update_button,
            self.open_records_button,
            self.delete_record_button,
            self.delete_input_button,
            self.output_browse_button,
            self.backup_check,
        ):
            button.config(state=state)
        self.input_text.config(state=state)
        self.output_entry.config(state=state)
        self.file_menu.entryconfig("选择输出目录…", state=state)
        self.file_menu.entryconfig("从已有文档导入链接…", state=state)
        self.edit_menu.entryconfig("清空输入", state=state)
        self.edit_menu.entryconfig("删除当前输入链接", state=state)
        self.edit_menu.entryconfig("删除选中记录", state=state)
        retry_state = "normal" if enabled and self.failed_jobs else "disabled"
        self.edit_menu.entryconfig("重试失败项", state=retry_state)
        self.stop_button.config(state="disabled" if enabled else "normal")

    def _build_ui(self) -> None:
        self.root.title("抖音信息提取工具")
        self.root.geometry("1120x680")
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        menu_bar = tk.Menu(self.root)
        self.file_menu = tk.Menu(menu_bar, tearoff=False)
        self.file_menu.add_command(label="选择输出目录…", command=self.browse_output)
        self.file_menu.add_command(
            label="从已有文档导入链接…", command=self.import_links_from_workbook
        )
        self.file_menu.add_separator()
        self.file_menu.add_command(label="打开输出目录", command=self.open_output)
        self.file_menu.add_command(label="打开封面文件夹", command=self.open_covers)
        self.file_menu.add_command(label="打开日志", command=self.open_log)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="退出", command=self._on_close)
        menu_bar.add_cascade(label="文件", menu=self.file_menu)

        self.edit_menu = tk.Menu(menu_bar, tearoff=False)
        self.edit_menu.add_command(label="清空输入", command=self.clear_input)
        self.edit_menu.add_command(label="删除当前输入链接", command=self.delete_current_input_link)
        self.edit_menu.add_command(label="复制选中标题", command=self.copy_title)
        self.edit_menu.add_command(label="删除选中记录", command=self.delete_selected_record)
        self.edit_menu.add_separator()
        self.edit_menu.add_command(label="重试失败项", command=self.retry_failed, state="disabled")
        menu_bar.add_cascade(label="编辑", menu=self.edit_menu)
        self.root.config(menu=menu_bar)

        main = ttk.Frame(self.root, padding=10)
        main.pack(fill="both", expand=True)

        input_header = ttk.Frame(main)
        input_header.pack(fill="x")
        ttk.Label(
            input_header,
            text="粘贴抖音链接（序号 1. 2. 3. … 已锁定不可修改，只修改序号后面的原始链接/分享文案；"
            "粘贴完自动弹出下一个序号）:",
        ).pack(side="left", anchor="w")
        self.delete_input_button = ttk.Button(
            input_header, text="删除当前链接", command=self.delete_current_input_link
        )
        self.delete_input_button.pack(side="right")

        input_frame = ttk.Frame(main)
        input_frame.pack(fill="x", pady=(4, 8))
        self.input_text = tk.Text(
            input_frame, height=7, wrap="word", font=("Microsoft YaHei UI", 10)
        )
        scroll = ttk.Scrollbar(input_frame, orient="vertical", command=self.input_text.yview)
        self.input_text.configure(yscrollcommand=scroll.set)
        self.input_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        # 自动编号：默认 1. ，粘贴完毕后自动弹出下一个序号，条目间用分割线分隔；
        # 序号在每次输入变化后按位置强制恢复，用户只能改后面的原始链接。
        self._renumber_after_id = None
        self.input_text.tag_configure("seq", foreground="#8a8a8a", background="#f4f4f4")
        self.input_text.tag_configure("divider", foreground="#b0b0b0")
        self.input_text.insert("1.0", "1. ")
        self.input_text.bind("<KeyRelease>", self._schedule_renumber)
        self.input_text.bind("<<Paste>>", self._on_input_paste)
        self.input_text.bind("<<Cut>>", self._on_input_cut)
        self.input_text.bind("<KeyPress-BackSpace>", self._on_input_delete_key)
        self.input_text.bind("<KeyPress-Delete>", self._on_input_delete_key)
        self.input_text.bind("<Control-Delete>", self.delete_current_input_link)

        output_frame = ttk.Frame(main)
        output_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(output_frame, text="输出目录:").pack(side="left")
        self.output_var = tk.StringVar(value=str(self.output_dir))
        self.output_entry = ttk.Entry(output_frame, textvariable=self.output_var)
        self.output_entry.pack(side="left", fill="x", expand=True, padx=(6, 6))
        self.output_browse_button = ttk.Button(
            output_frame, text="选择目录…", command=self.browse_output
        )
        self.output_browse_button.pack(side="right")
        self.backup_var = tk.BooleanVar(value=False)
        self.backup_check = ttk.Checkbutton(
            output_frame,
            text="备份文件",
            variable=self.backup_var,
        )
        self.backup_check.pack(side="right", padx=(4, 8))

        button_frame = ttk.Frame(main)
        button_frame.pack(fill="x", pady=(0, 8))
        for column in range(5):
            button_frame.columnconfigure(column, weight=1, uniform="main_actions")
        self.start_button = ttk.Button(button_frame, text="全部提取", command=self.start)
        self.start_button.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        self.single_button = ttk.Button(
            button_frame, text="单次提取", command=self.extract_current_once
        )
        self.single_button.grid(row=0, column=1, sticky="ew", padx=4)
        self.update_button = ttk.Button(
            button_frame, text="刷新记录", command=self.refresh_table_data
        )
        self.update_button.grid(row=0, column=2, sticky="ew", padx=4)
        self.open_records_button = ttk.Button(
            button_frame, text="打开提取记录", command=self.open_records
        )
        self.open_records_button.grid(row=0, column=3, sticky="ew", padx=4)
        self.delete_record_button = ttk.Button(
            button_frame, text="删除选中记录", command=self.delete_selected_record
        )
        self.delete_record_button.grid(row=0, column=4, sticky="ew", padx=(4, 0))

        progress_frame = ttk.Frame(main)
        progress_frame.pack(fill="x", pady=(0, 8))
        self.progress = ttk.Progressbar(progress_frame, mode="determinate", maximum=100)
        self.progress.pack(side="left", fill="x", expand=True)
        self.progress_label = ttk.Label(progress_frame, text="", width=30, anchor="e")
        self.progress_label.pack(side="left", padx=(6, 0))
        self.stop_button = ttk.Button(
            progress_frame, text="停止任务", command=self.stop_task, state="disabled"
        )
        self.stop_button.pack(side="right", padx=(6, 0))

        result_frame = ttk.Frame(main)
        result_frame.pack(fill="both", expand=True)

        columns = ("title", "likes", "comments", "media", "status")
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings")
        self.tree.tag_configure(
            "refreshing", background="#e8f5e9", foreground="#1b5e20"
        )
        self.tree.tag_configure(
            "refresh_success", background="#d7f2df", foreground="#145c2e"
        )
        self.tree.tag_configure(
            "refresh_failure", background="#fde0e0", foreground="#9b1c1c"
        )
        headings = {
            "title": ("标题", 300),
            "likes": ("点赞数", 90),
            "comments": ("评论数", 90),
            "media": ("视频/图集", 140),
            "status": ("状态", 240),
        }
        for column, (text, width) in headings.items():
            self.tree.heading(column, text=text)
            self.tree.column(column, width=width, anchor="w")
        self.tree.pack(side="left", fill="both", expand=True)

        tree_scroll = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        tree_scroll.pack(side="left", fill="y")
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Button-3>", self._show_record_menu)
        self.tree.bind("<Delete>", self.delete_selected_record)
        self.record_menu = tk.Menu(self.tree, tearoff=False)
        self.record_menu.add_command(label="删除选中记录", command=self.delete_selected_record)

        self.preview_frame = ttk.Frame(
            result_frame,
            width=PREVIEW_BOX_SIZE[0],
            height=PREVIEW_BOX_SIZE[1],
        )
        self.preview_frame.pack(side="left", anchor="n", padx=(8, 0))
        self.preview_frame.pack_propagate(False)
        self.preview = tk.Label(
            self.preview_frame,
            text="选中一行查看封面预览",
            bg="#f2f2f2",
            relief="groove",
            anchor="center",
            justify="center",
        )
        self.preview.pack(fill="both", expand=True)

        self.status_var = tk.StringVar(
            value="就绪：粘贴链接，选择输出目录后点击「全部提取」"
            "（右键下方记录可同步删除链接、表格行和关联文件）"
        )
        ttk.Label(main, textvariable=self.status_var, anchor="w").pack(fill="x", pady=(8, 0))

    def load_existing_records(self) -> None:
        """从输出目录的表格与文件加载已有记录，显示到结果列表（下次打开也能看到）。"""
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        self.records.clear()
        output_dir = Path(self.output_var.get().strip() or default_output_dir())
        try:
            rows = exporter.read_records(output_dir / "提取记录.xlsx")
        except Exception as exc:
            self.status_var.set(f"记录表无法读取：{exc}（原文件未修改）")
            return
        videos_dir = output_dir / "爆款视频"
        cover_dir = output_dir / "封面"
        for seq in sorted(rows):
            rec = rows[seq]
            media_display = "未下载"
            if videos_dir.exists():
                video_file = videos_dir / f"{seq}.mp4"
                image_dir = videos_dir / str(seq)
                if video_file.exists():
                    media_display = f"{seq}.mp4"
                elif image_dir.is_dir():
                    count = sum(1 for f in image_dir.iterdir() if f.is_file())
                    media_display = f"{seq}/（{count} 张图）"
            cover_path = None
            if cover_dir.exists():
                matches = sorted(cover_dir.glob(f"{seq}.*"))
                if matches:
                    cover_path = str(matches[0])
            status = rec.get("status") or ""
            # 绿色只表示本轮刷新刚刚成功；任务结束重新加载后，正常/已恢复
            # 都回到无色，只有异常状态常驻红色。
            row_tag = "" if status in {"", "正常", "已恢复"} else "refresh_failure"
            item_id = self.tree.insert(
                "",
                "end",
                values=(
                    rec.get("title") or "无",
                    f"{rec.get('likes') or 0:,}",
                    f"{rec.get('comments') or 0:,}",
                    media_display,
                    status or "—",
                ),
                tags=(row_tag,) if row_tag else (),
            )
            self.records[item_id] = {
                "title": rec.get("title") or "",
                "likes": rec.get("likes") or 0,
                "comments": rec.get("comments") or 0,
                "media_display": media_display,
                "cover_path": cover_path,
                "seq": seq,
                "raw_input": rec.get("raw_input") or "",
                "status": rec.get("status") or "",
            }

    def _highlight_record(self, seq: int, tag: str) -> None:
        """按表格顺序号高亮刷新中的记录及其最终结果。"""
        for item_id, record in self.records.items():
            if record.get("seq") == seq:
                self.tree.item(item_id, tags=(tag,))
                self.tree.see(item_id)
                return

    def _update_visible_record(self, seq: int, record: dict, tag: str) -> None:
        """实时更新已检查行的最新数据、状态和结果颜色。"""
        for item_id, visible in self.records.items():
            if visible.get("seq") != seq:
                continue
            visible.update(
                {
                    "title": record.get("title") or "",
                    "likes": record.get("likes") or 0,
                    "comments": record.get("comments") or 0,
                    "status": record.get("status") or "",
                }
            )
            self.tree.item(
                item_id,
                values=(
                    visible["title"] or "无",
                    f"{visible['likes']:,}",
                    f"{visible['comments']:,}",
                    visible.get("media_display") or "未下载",
                    visible["status"] or "—",
                ),
                tags=(tag,),
            )
            self.tree.see(item_id)
            return

    def _style_input_sequences(self) -> None:
        """把锁定序号与分割线标灰，提示这两部分不可编辑（内容可编辑）。"""
        widget = self.input_text
        widget.tag_remove("seq", "1.0", "end")
        widget.tag_remove("divider", "1.0", "end")
        expected = 0
        at_block_start = True
        for line_no, line in enumerate(widget.get("1.0", "end").splitlines(), 1):
            stripped = line.strip()
            if DIVIDER_RE.match(stripped):
                widget.tag_add("divider", f"{line_no}.0", f"{line_no}.end")
                at_block_start = True
            elif stripped and at_block_start:
                expected += 1
                prefix = f"{expected}."
                widget.tag_add("seq", f"{line_no}.0", f"{line_no}.0 + {len(prefix)} chars")
                at_block_start = False

    def _enforce_input_sequences(self) -> None:
        """强制序号 = 条目位置。

        无论用户怎么改序号，松手/粘贴后都会恢复为 1. 2. 3. …；
        原始链接/分享文案原样保留。这样第 5 条就是 5.，不会再被改成 9.
        而导致后续识别错位。
        """
        widget = self.input_text
        current = widget.get("1.0", "end").rstrip("\n")
        normalized = normalize_input_text(current)
        if normalized == current:
            self._style_input_sequences()
            return
        widget.delete("1.0", "end")
        widget.insert("1.0", normalized)
        widget.mark_set("insert", "end-1c")
        widget.see("insert")
        self._style_input_sequences()
        self._save_input_cache()

    @staticmethod
    def _build_existing_link_map(rows: dict[int, dict]) -> dict[str, int]:
        """表格已有链接 → 顺序。"""
        link_map: dict[str, int] = {}
        for seq, rec in rows.items():
            for url in extractor.extract_urls(rec.get("raw_input") or ""):
                link_map.setdefault(url, seq)
        return link_map

    def _existing_media_targets(
        self, jobs: list[tuple[int | None, str]], output_dir: Path, rows: dict[int, dict]
    ) -> list[tuple[int, str]]:
        """找出输入里已提取过且视频/图集还在的链接，返回 [(顺序, 原始链接)]。"""
        link_map = self._build_existing_link_map(rows)
        videos_dir = output_dir / "爆款视频"
        targets: list[tuple[int, str]] = []
        seen: set[int] = set()
        for _seq, raw in jobs:
            for url in extractor.extract_urls(raw):
                hit_seq = link_map.get(url)
                if not hit_seq or hit_seq in seen:
                    continue
                if (videos_dir / f"{hit_seq}.mp4").exists() or (
                    videos_dir / str(hit_seq)
                ).is_dir():
                    targets.append((hit_seq, raw))
                    seen.add(hit_seq)
                    break
        return targets

    def browse_output(self) -> None:
        chosen = filedialog.askdirectory(
            title="选择输出目录", initialdir=self.output_var.get() or default_output_dir()
        )
        if not chosen:
            return
        self.output_var.set(chosen)
        self.output_dir = Path(chosen)
        save_config({"output_dir": str(self.output_dir)})
        self.load_existing_records()

    def import_links_from_workbook(self) -> None:
        """只读旧提取记录，按「顺序」把其中的纯链接回填到输入框。"""
        if self.running or self.refreshing:
            return
        initial_dir = self.output_var.get().strip() or str(default_output_dir())
        chosen = filedialog.askopenfilename(
            title="选择以前的提取记录文档",
            initialdir=initial_dir,
            filetypes=(
                ("Excel 工作簿", ("*.xlsx", "*.xlsm")),
                ("所有文件", "*.*"),
            ),
        )
        if not chosen:
            return

        path = Path(chosen)
        logger = logging.getLogger("douyin_tool")
        try:
            rows = exporter.read_records(path)
        except Exception as exc:
            logger.warning("导入旧文档失败：%s：%s", path.name, exc)
            messagebox.showerror(
                "无法读取文档",
                f"无法读取“{path.name}”：\n{exc}\n\n原文档和当前输入均未修改。",
                parent=self.root,
            )
            return

        links = input_parser.links_from_records_in_sequence(rows)
        if not links:
            messagebox.showinfo(
                "没有可导入的链接",
                "文档中没有同时具备有效“顺序”和抖音链接的记录。\n"
                "请选择本工具以前生成的提取记录.xlsx。",
                parent=self.root,
            )
            return

        current_links = extractor.extract_urls(self.input_text.get("1.0", "end"))
        if current_links and not messagebox.askyesno(
            "替换当前输入",
            f"当前输入框已有 {len(current_links)} 条链接。\n"
            f"是否替换为“{path.name}”中的 {len(links)} 条链接？",
            parent=self.root,
        ):
            self.status_var.set("已取消导入，当前输入未修改")
            return

        formatted = input_parser.format_ordered_links(links)
        self.input_text.delete("1.0", "end")
        self.input_text.insert("1.0", formatted)
        self.input_text.mark_set("insert", "end-1c")
        self.input_text.see("insert")
        self._style_input_sequences()
        self._save_input_cache()
        self.status_var.set(
            f"已从“{path.name}”按文档顺序导入 {len(links)} 条链接（未联网、未修改原文档）"
        )
        logger.info("从旧文档导入链接：%s，共 %d 条", path.name, len(links))

    def start(self, jobs_override: list[tuple[int | None, str]] | None = None) -> None:
        if self.running or self.refreshing:
            return
        retry_run = bool(jobs_override is not None and self._retry_request)
        self._retry_request = False
        single_run = jobs_override is not None and not retry_run
        if jobs_override is not None:
            jobs = list(jobs_override or [])
            ignored = 0
        else:
            self._enforce_input_sequences()
            jobs, ignored = build_input_jobs(self.input_text.get("1.0", "end"))
        if not jobs:
            self.status_var.set("请先粘贴至少一条抖音链接")
            return
        self._save_input_cache()

        output_dir = Path(self.output_var.get().strip() or default_output_dir())
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            (output_dir / "爆款视频").mkdir(parents=True, exist_ok=True)
            (output_dir / "文案提取").mkdir(parents=True, exist_ok=True)
            (output_dir / "封面").mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            self.status_var.set(f"输出目录不可用：{exc}")
            return
        self.output_dir = output_dir
        save_config({"output_dir": str(output_dir)})
        self.logger = setup_logger(output_dir / LOG_NAME)
        self.logger.info("=" * 60)
        action_name = "重试失败项" if retry_run else ("单次提取" if single_run else "全部提取")
        self.logger.info("开始%s：共 %d 条，输出目录 %s", action_name, len(jobs), output_dir)
        self.backup_enabled = bool(self.backup_var.get())
        self.logger.info(
            "长期备份：%s（失败回滚临时文件不受此选项影响）",
            "开启" if self.backup_enabled else "关闭",
        )
        if not retry_run:
            self.logger.info("表格已有记录只强制刷新最新公开数据并原子写回 Excel，绝不下载媒体")

        # —— 多重检测 1：提取前检查编号一致性，能自动修的立即修，需确认的弹窗提示 ——
        audit_note = ""
        videos_n, captions_n, covers_n = scan_seq_numbers(output_dir)
        note_n = scan_note_numbers(output_dir)
        # 图文（图集目录）不生成文案文件，只给视频补建空文案
        for n in sorted((videos_n - note_n) - captions_n):
            try:
                (output_dir / "文案提取" / f"{n}.txt").write_text("", encoding="utf-8")
                self.logger.warning("顺序检测：序号 %d 有视频缺文案，已自动补建空文案文件", n)
            except OSError as exc:
                self.logger.error("顺序检测：补建 %d.txt 失败：%s", n, exc)
        orphan_captions = captions_n - videos_n
        orphan_covers = covers_n - videos_n
        excel_mismatch = _check_excel_order(output_dir, videos_n)
        if orphan_captions or orphan_covers or excel_mismatch:
            notes = [
                f"· 序号 {n}：只有文案文件 {n}.txt，没有对应视频" for n in sorted(orphan_captions)
            ]
            notes += [
                f"· 序号 {n}：只有封面文件，没有对应视频" for n in sorted(orphan_covers)
            ]
            if excel_mismatch:
                notes.append("· 表格「顺序」列与文件编号不一致，请手动核对")
            msg = (
                "提取前检测到顺序问题：\n"
                + "\n".join(notes)
                + "\n\n多余文件是否移入「顺序修复备份」文件夹以修正顺序？"
                "\n（选「否」保留原样，编号沿用旧最大值）"
            )
            if messagebox.askyesno("顺序检测", msg, parent=self.root):
                backup_dir = output_dir / "顺序修复备份" / time.strftime("%Y-%m-%d_%H%M%S")
                backup_dir.mkdir(parents=True, exist_ok=True)
                for n in sorted(orphan_captions):
                    src = output_dir / "文案提取" / f"{n}.txt"
                    if src.exists():
                        shutil.move(str(src), str(backup_dir / f"{n}.txt"))
                for n in sorted(orphan_covers):
                    for src in (output_dir / "封面").glob(f"{n}.*"):
                        shutil.move(str(src), str(backup_dir / src.name))
                self.logger.warning(
                    "已按用户确认将多余文件移入「顺序修复备份」（文案 %d 个、封面 %d 个）",
                    len(orphan_captions),
                    len(orphan_covers),
                )
                audit_note = "，多余文件已移入「顺序修复备份」"
            else:
                self.logger.warning("检测到顺序问题但用户选择保留原样")
                audit_note = "，⚠ 顺序问题未处理"

        self.running = True
        self.single_run = single_run
        self.retry_run = retry_run
        self.cancel_event.clear()
        self.close_requested = False
        self.ok_count = 0
        self.fail_count = 0
        self.refresh_count = 0
        self.cancel_count = 0
        self.rollback_count = 0
        self.unchecked_count = 0
        self.pause_message = ""
        self.failed_jobs = []

        self._set_task_controls(False)
        if single_run:
            self.single_button.config(text="提取中…")
        else:
            self.start_button.config(text="提取中…")
        ignore_note = f"，忽略 {ignored} 行无链接内容" if ignored else ""
        if single_run:
            self.status_var.set(f"正在单次处理；表格已有记录只强制刷新 Excel{audit_note}…")
        elif retry_run:
            self.status_var.set(f"正在重试 {len(jobs)} 条失败项{audit_note}…")
        else:
            self.status_var.set(
                f"开始全部处理，共 {len(jobs)} 条；表格已有记录只强制刷新 Excel"
                f"{audit_note}{ignore_note}…"
            )
        self.worker = threading.Thread(
            target=self._work_safe,
            args=(jobs, output_dir),
            daemon=False,
        )
        self.worker.start()
        self.root.after(100, self._poll)

    def _work_safe(
        self,
        jobs: list[tuple[int | None, str]],
        output_dir: Path,
    ) -> None:
        """带取消、稳定 ID 去重和文件事务的正式提取流程。"""
        logger = logging.getLogger("douyin_tool")
        videos_dir = output_dir / "爆款视频"
        xlsx = output_dir / "提取记录.xlsx"
        current_job: tuple[int | None, str] | None = None
        access_context = extractor.AccessContext(
            BROWSER_PROFILE_DIR,
            self.cancel_event,
            lambda event, message: self._post(
                "verification", {"event": event}, message
            ),
        )

        def has_media(seq: int) -> bool:
            return (videos_dir / f"{seq}.mp4").is_file() or (videos_dir / str(seq)).is_dir()

        try:
            existing_rows = exporter.read_records(xlsx)
            link_map = self._build_existing_link_map(existing_rows)
            id_map = {
                str(rec.get("aweme_id")): seq
                for seq, rec in existing_rows.items()
                if str(rec.get("aweme_id") or "").strip()
            }
            used_seqs = set(existing_rows)
            used_seqs.update(scan_seq_numbers(output_dir)[0])

            normalized_jobs = (
                [(None, line) for line in jobs] if jobs and isinstance(jobs[0], str) else jobs
            )
            for index, current_job in enumerate(normalized_jobs, 1):
                ensure_not_cancelled(self.cancel_event)
                input_seq, line = current_job
                if index > 1:
                    interruptible_wait(1 + random.random(), self.cancel_event)
                seq_hint = input_seq if input_seq is not None else ((max(used_seqs) + 1) if used_seqs else 1)
                logger.info("[%d/%d] 抓取顺序候选 %d：%s", index, len(normalized_jobs), seq_hint, line[:80])

                exact_hit = next(
                    (
                        link_map[url]
                        for url in extractor.extract_urls(line)
                        if url in link_map
                    ),
                    None,
                )

                pause_reason = ""
                try:
                    fetched, fail_status = fetch_with_retry(
                        logger,
                        seq_hint,
                        line,
                        self.cancel_event,
                        access_context,
                    )
                except extractor.BrowserVerificationError as exc:
                    fetched, fail_status = None, exc.status
                    pause_reason = str(exc)
                if fetched is None:
                    # 已有记录本次抓取失败时保留标题、互动数和人工字段，但状态
                    # 必须反映本次检查结果，不能继续冒充“正常”。
                    if exact_hit is not None:
                        failed_record = dict(existing_rows.get(exact_hit) or {})
                        failed_record["status"] = fail_status
                        failed_record["updated_at"] = datetime.now().strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        try:
                            update_records_force_close(
                                xlsx,
                                {exact_hit: failed_record},
                                [exact_hit],
                                keep_backup=getattr(self, "backup_enabled", False),
                            )
                            existing_rows[exact_hit] = failed_record
                        except Exception as exc:
                            logger.error("顺序 %d 失败状态写回失败：%s", exact_hit, exc)
                            fail_status = f"{fail_status}；状态写回失败（{exc}）"
                    self._post(
                        "error",
                        {"input": line, "job": current_job, "seq": exact_hit or seq_hint},
                        fail_status,
                    )
                    if pause_reason:
                        unchecked = len(normalized_jobs) - index
                        logger.warning(
                            "浏览器验证未完成，批次熔断：当前失败 1 条，未检查 %d 条",
                            unchecked,
                        )
                        self._post(
                            "paused",
                            {
                                "job": current_job,
                                "unchecked": unchecked,
                            },
                            f"批次已暂停：{pause_reason}",
                        )
                        break
                    continue

                id_hit = id_map.get(fetched.aweme_id)
                hit_seq = id_hit or exact_hit

                # 恢复旧版分流：只要链接或作品 ID 已在表格中，就永远只刷新
                # Excel，不以媒体是否存在为条件，也不补下载、不替换媒体。
                if hit_seq is not None:
                    updated = dict(existing_rows.get(hit_seq) or {})
                    updated.update(
                        {
                            "raw_input": line,
                            "title": fetched.fields["title"],
                            "tags": fetched.fields["tags"],
                            "likes": fetched.fields["likes"],
                            "comments": fetched.fields["comments"],
                            "author": fetched.fields["author"],
                            "status": "正常",
                            "aweme_id": fetched.aweme_id,
                            "work_kind": "图文" if fetched.kind == "note" else "视频",
                            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                    )
                    try:
                        update_records_force_close(
                            xlsx,
                            {hit_seq: updated},
                            [hit_seq],
                            keep_backup=getattr(self, "backup_enabled", False),
                        )
                    except Exception as exc:
                        logger.error("顺序 %d 元数据写回失败：%s", hit_seq, exc)
                        self._post(
                            "error",
                            {"input": line, "job": current_job, "seq": hit_seq},
                            str(exc),
                        )
                        continue
                    existing_rows[hit_seq] = updated
                    id_map[fetched.aweme_id] = hit_seq
                    for url in extractor.extract_urls(line):
                        link_map[url] = hit_seq
                    logger.info(
                        "作品 %s 已有表格记录，顺序 %d 已强制刷新 Excel，未进入媒体下载",
                        fetched.aweme_id,
                        hit_seq,
                    )
                    self._post(
                        "refreshed",
                        {"seq": hit_seq, "reason": "已有记录已强制刷新 Excel，未下载媒体"},
                    )
                    continue

                if hit_seq is not None:
                    seq = hit_seq
                elif input_seq is not None:
                    seq = input_seq
                else:
                    seq = (max(used_seqs) + 1) if used_seqs else 1
                replacing = seq in existing_rows or has_media(seq)
                transaction = ArtifactTransaction(
                    output_dir,
                    seq,
                    keep_backup=getattr(self, "backup_enabled", False),
                )
                try:
                    if fetched.kind == "note":
                        staged_media = transaction.note_target()

                        def image_progress(done, total, n=seq):
                            self._post(
                                "progress",
                                {"seq": n, "done": done, "total": total, "unit": "images"},
                            )

                        paths = extractor.download_images(
                            fetched.session,
                            fetched.item,
                            staged_media,
                            image_progress,
                            self.cancel_event,
                            browser_context=access_context.browser_context,
                            browser_context_provider=access_context.ensure_browser_context,
                        )
                        hits = [
                            find_same_size_file(videos_dir, path.stat().st_size)
                            for path in paths
                        ]
                        if paths and all(hits):
                            logger.warning(
                                "作品 %s 的图集大小与旧文件相似，仅记录提醒，不自动判重",
                                fetched.aweme_id,
                            )
                        media_display = f"{seq}/（{len(paths)} 张图）"
                    else:
                        staged_media = transaction.video_target()

                        def video_progress(done, total, n=seq):
                            self._post(
                                "progress",
                                {"seq": n, "done": done, "total": total, "unit": "bytes"},
                            )

                        # 浏览器流在响应头到达前没有字节可回调，先明确显示当前
                        # 已进入媒体下载阶段，避免用户误以为提取或写入线程卡死。
                        video_progress(0, 0)
                        extractor.download_video(
                            fetched.session,
                            fetched.item,
                            staged_media,
                            video_progress,
                            cancel_event=self.cancel_event,
                            browser_context=access_context.browser_context,
                            browser_context_provider=access_context.ensure_browser_context,
                        )
                        size_hit = find_same_size_file(
                            videos_dir, staged_media.stat().st_size
                        )
                        if size_hit:
                            logger.warning(
                                "作品 %s 与 %s 大小相同，仅记录提醒，不自动判重",
                                fetched.aweme_id,
                                size_hit,
                            )
                        media_display = f"{seq}.mp4"

                    staged_cover = None
                    if fetched.fields.get("cover_url"):
                        try:
                            staged_cover = extractor.download_cover(
                                fetched.session,
                                fetched.fields["cover_url"],
                                transaction.cover_dir(),
                                str(seq),
                                self.cancel_event,
                                browser_context=access_context.browser_context,
                                browser_context_provider=access_context.ensure_browser_context,
                            )
                        except TaskCancelled:
                            raise
                        except Exception:
                            try:
                                fresh_session, fresh_item = extractor.fetch_item_with_session(
                                    fetched.session,
                                    fetched.aweme_id,
                                    fetched.kind,
                                    self.cancel_event,
                                )
                            except extractor.NetworkRequestError:
                                # 封面地址过期且 Requests 链路仍不可用时，
                                # 复用当前浏览器上下文重新取得作品数据。
                                fresh = access_context.fetch_record(line)
                                fresh_session, fresh_item = fresh.session, fresh.item
                            fresh_fields = extractor.extract_fields(
                                fresh_item, fetched.aweme_id
                            )
                            if not fresh_fields.get("cover_url"):
                                raise extractor.ExtractionError("封面地址不可用")
                            staged_cover = extractor.download_cover(
                                fresh_session,
                                fresh_fields["cover_url"],
                                transaction.cover_dir(),
                                str(seq),
                                self.cancel_event,
                                browser_context=access_context.browser_context,
                                browser_context_provider=access_context.ensure_browser_context,
                            )

                    old_record = existing_rows.get(seq) or {}
                    record = {
                        "raw_input": line,
                        "title": fetched.fields["title"],
                        "tags": fetched.fields["tags"],
                        "likes": fetched.fields["likes"],
                        "comments": fetched.fields["comments"],
                        "type": old_record.get("type") or "基本盘",
                        "author": fetched.fields["author"],
                        "status": "正常",
                        "aweme_id": fetched.aweme_id,
                        "work_kind": "图文" if fetched.kind == "note" else "视频",
                        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "media_display": media_display,
                        "seq": seq,
                        "replace": replacing,
                    }

                    def persist_workbook(final_cover: Path | None) -> None:
                        latest_rows = exporter.read_records(xlsx)
                        latest_manual = latest_rows.get(seq) or {}
                        if latest_manual.get("type"):
                            record["type"] = latest_manual["type"]
                        latest_rows[seq] = record
                        cover_map = build_cover_map(output_dir, set(latest_rows))
                        cover_map[seq] = str(final_cover) if final_cover else None
                        update_records_force_close(
                            xlsx,
                            latest_rows,
                            [seq],
                            cover_map,
                            keep_backup=getattr(self, "backup_enabled", False),
                        )

                    result = transaction.commit(
                        work_kind=fetched.kind,
                        staged_media=staged_media,
                        staged_cover=staged_cover,
                        persist_workbook=persist_workbook,
                    )
                    record["cover_path"] = str(result.cover_path) if result.cover_path else None
                    existing_rows[seq] = dict(record)
                    used_seqs.add(seq)
                    id_map[fetched.aweme_id] = seq
                    for url in extractor.extract_urls(line):
                        link_map[url] = seq
                    logger.info("顺序 %d 成功，作品 ID %s：%s", seq, fetched.aweme_id, fetched.fields["title"][:40])
                    self._post("ok", record, "")
                except TaskCancelled:
                    transaction.cleanup()
                    raise
                except TransactionError as exc:
                    transaction.cleanup()
                    self._post(
                        "error",
                        {
                            "input": line,
                            "job": current_job,
                            "seq": seq,
                            "rolled_back": exc.rolled_back,
                        },
                        str(exc),
                    )
                except Exception as exc:
                    transaction.cleanup()
                    logger.error("顺序 %d 处理失败：%s", seq, exc)
                    self._post(
                        "error",
                        {"input": line, "job": current_job, "seq": seq},
                        str(exc),
                    )
        except TaskCancelled:
            logger.info("任务已按用户请求取消")
            self._post("cancelled", {"job": current_job})
        except Exception as exc:
            logger.exception("提取任务异常终止")
            self._post(
                "fatal",
                {"job": current_job},
                f"任务异常终止，原有文件未主动修改：{exc}",
            )
        finally:
            access_context.close()
            self._post("done")

    def _poll(self) -> None:
        try:
            while True:
                kind, payload, extra = self.message_queue.get_nowait()
                if kind == "ok":
                    self.ok_count += 1
                    self.progress["value"] = 100
                    self.progress_label.config(text=f"{payload['seq']} 完成")
                    # 原位替换不往页面重复插行，批次结束后统一重载表格
                    if not payload.get("replace"):
                        item_id = self.tree.insert(
                            "",
                            "end",
                            values=(
                                payload["title"],
                                f"{payload['likes']:,}",
                                f"{payload['comments']:,}",
                                payload["media_display"],
                                "成功" + (extra or ""),
                            ),
                        )
                        self.records[item_id] = payload
                    continue
                if kind == "refreshed":
                    self.refresh_count += 1
                    self.status_var.set(
                        f"序号 {payload.get('seq')}："
                        f"{payload.get('reason') or '已有记录已强制刷新 Excel'}"
                    )
                    continue
                if kind == "verification":
                    self.status_var.set(extra or "请在弹出的浏览器中完成验证")
                    event = (payload or {}).get("event") if isinstance(payload, dict) else ""
                    self.progress_label.config(
                        text=(
                            "网络链路切换中…"
                            if event in {"network_fallback", "network_retry"}
                            else "等待浏览器验证…"
                        )
                    )
                    continue
                if kind == "paused":
                    self.unchecked_count = int(payload.get("unchecked") or 0)
                    self.pause_message = extra or "批次已暂停"
                    self.status_var.set(self.pause_message)
                    continue
                if kind == "dup_status":
                    self.refresh_count += payload.get("total") or 0
                    self.status_var.set(
                        f"已有记录已强制刷新 Excel {payload.get('total') or 0} 条，"
                        "未进入媒体下载…"
                    )
                    continue
                if kind == "progress":
                    total = payload.get("total") or 0
                    done = payload.get("done") or 0
                    percent = int(done / total * 100) if total else 0
                    self.progress["value"] = percent
                    if payload.get("unit") == "images":
                        text = f"下载图集 {payload['seq']}：{done}/{total} 张"
                    elif total:
                        text = f"下载视频 {payload['seq']}：{done // 1048576}MB/{total // 1048576}MB"
                    else:
                        text = f"下载视频 {payload['seq']}：{done // 1048576}MB"
                    self.progress_label.config(text=text)
                elif kind == "error":
                    self.fail_count += 1
                    if payload.get("rolled_back"):
                        self.rollback_count += 1
                    job = payload.get("job")
                    if job and job not in self.failed_jobs:
                        self.failed_jobs.append(job)
                    input_text = payload.get("input") or ""
                    item_id = self.tree.insert(
                        "", "end", values=(input_text[:40], "", "", "", extra)
                    )
                    self.records[item_id] = {"title": "", "input": input_text, "message": extra}
                elif kind == "cancelled":
                    self.cancel_count += 1
                    self.status_var.set("任务已取消，已完成当前清理")
                elif kind == "fatal":
                    self.fail_count += 1
                    job = payload.get("job") if isinstance(payload, dict) else None
                    if job and job not in self.failed_jobs:
                        self.failed_jobs.append(job)
                    self.status_var.set(extra or "任务异常终止")
                elif kind == "done":
                    self._finish()
                    return
        except queue.Empty:
            pass
        if self.running:
            self.root.after(100, self._poll)

    def _finish(self) -> None:
        self.running = False
        self._set_task_controls(True)
        self.start_button.config(text="全部提取")
        self.single_button.config(text="单次提取")
        self.progress["value"] = 0
        self.progress_label.config(text="")
        self.load_existing_records()
        prefix = (
            "失败项重试完成"
            if getattr(self, "retry_run", False)
            else ("单次提取完成" if getattr(self, "single_run", False) else "完成")
        )
        summary = (
            f"{prefix}：新提取/替换 {self.ok_count}，强制刷新写表 {self.refresh_count}，"
            f"失败 {self.fail_count}，未检查 {self.unchecked_count}，"
            f"取消 {self.cancel_count}，已回滚 {self.rollback_count}"
        )
        if self.pause_message:
            summary = f"{summary}；{self.pause_message}"
        self.status_var.set(summary)
        logger = logging.getLogger("douyin_tool")
        logger.info(
            "批次完成：新提取/替换 %d，强制刷新写表 %d，失败 %d，未检查 %d，取消 %d，已回滚 %d",
            self.ok_count,
            self.refresh_count,
            self.fail_count,
            self.unchecked_count,
            self.cancel_count,
            self.rollback_count,
        )
        # —— 多重检测 3：批次结束后复核顺序一致性 ——
        try:
            videos_n, captions_n, covers_n = scan_seq_numbers(self.output_dir)
            note_n = scan_note_numbers(self.output_dir)
            video_seqs = videos_n - note_n
            if video_seqs - captions_n:
                logger.warning("批次后复核：序号 %s 缺文案文件", sorted(video_seqs - captions_n))
            if captions_n - videos_n:
                logger.warning("批次后复核：序号 %s 只有文案没有视频", sorted(captions_n - videos_n))
            if covers_n - videos_n:
                logger.warning("批次后复核：序号 %s 只有封面没有视频", sorted(covers_n - videos_n))
        except Exception as exc:
            logger.warning("批次后复核失败：%s", exc)
        if self.close_requested:
            self.root.after(50, self._wait_then_close)

    def _on_select(self, _event=None) -> None:
        selection = self.tree.selection()
        if not selection:
            return
        record = self.records.get(selection[0])
        cover_path = record.get("cover_path") if record else None
        try:
            image = prepare_preview_image(cover_path)
            self.thumb_ref = ImageTk.PhotoImage(image)
            self.preview.config(image=self.thumb_ref, text="")
        except Exception:
            self.thumb_ref = None
            self.preview.config(image="", text="封面不可用")

    def _show_record_menu(self, event) -> str:
        """右键选中光标下的记录并显示删除入口。"""
        item_id = self.tree.identify_row(event.y)
        if item_id:
            self.tree.selection_set(item_id)
            self.tree.focus(item_id)
        record = self.records.get(item_id) if item_id else None
        can_delete = bool(
            record
            and record.get("seq")
            and not self.running
            and not self.refreshing
        )
        self.record_menu.entryconfig(
            "删除选中记录", state="normal" if can_delete else "disabled"
        )
        try:
            self.record_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.record_menu.grab_release()
        return "break"

    def delete_selected_record(self, _event=None) -> str:
        """同步删除输入链接、Excel 行和该顺序关联文件，并重排后续顺序。"""
        if self.running or self.refreshing:
            self.status_var.set("正在提取/刷新中，不能删除记录")
            return "break"
        selection = self.tree.selection()
        if not selection:
            self.status_var.set("请先在下方选择一条已提取记录")
            return "break"
        item_id = selection[0]
        record = self.records.get(item_id) or {}
        seq = record.get("seq")
        if not isinstance(seq, int) or seq <= 0:
            self.status_var.set("这不是已经写入提取记录的项目，无法同步删除")
            return "break"

        keep_backup = bool(self.backup_var.get())
        title = str(record.get("title") or "无标题")
        backup_note = (
            "删除内容会移入“删除备份”。"
            if keep_backup
            else "“备份文件”未勾选，删除内容成功后不可恢复。"
        )
        confirmed = messagebox.askyesno(
            "删除提取记录",
            f"确定删除顺序 {seq}：{title[:50]}？\n\n"
            "将同步删除：\n"
            "· 上方输入框中的对应链接\n"
            "· 下方列表和提取记录.xlsx 对应行\n"
            "· 对应视频/图集、封面和文案\n"
            "· 后续记录及文件顺序将整体前移一位\n\n"
            f"{backup_note}",
            parent=self.root,
        )
        if not confirmed:
            return "break"

        output_dir = Path(self.output_var.get().strip() or default_output_dir())
        xlsx = output_dir / "提取记录.xlsx"
        logger = setup_logger(output_dir / LOG_NAME)
        old_input = self.input_text.get("1.0", "end")
        transaction = RecordDeletionTransaction(
            output_dir, seq, keep_backup=keep_backup
        )
        try:
            result = transaction.commit(
                lambda: delete_record_force_close(
                    xlsx, seq, keep_backup=keep_backup
                )
            )
        except TransactionError as exc:
            logger.error("删除顺序 %d 失败：%s", seq, exc)
            messagebox.showerror(
                "删除失败",
                str(exc),
                parent=self.root,
            )
            self.status_var.set(f"删除顺序 {seq} 失败，原记录已尽量恢复")
            return "break"
        except Exception as exc:
            logger.exception("删除顺序 %d 异常", seq)
            messagebox.showerror("删除失败", str(exc), parent=self.root)
            self.status_var.set(f"删除顺序 {seq} 失败：{exc}")
            return "break"

        new_input, removed_links = input_parser.remove_matching_entry(
            old_input, seq, str(record.get("raw_input") or "")
        )
        self.input_text.delete("1.0", "end")
        self.input_text.insert("1.0", new_input)
        self.input_text.mark_set("insert", "end-1c")
        self.input_text.see("insert")
        self._style_input_sequences()
        self._save_input_cache()
        self.thumb_ref = None
        self.preview.config(image="", text="选中一行查看封面预览")
        self.load_existing_records()
        logger.info(
            "已删除顺序 %d；输入链接 %d 条，关联文件 %d 个，后续重排文件 %d 个，长期备份 %s",
            seq,
            removed_links,
            result.deleted_artifacts,
            result.shifted_artifacts,
            "有" if result.backup_dir else "无",
        )
        link_note = "" if removed_links else "（上方未找到对应链接）"
        self.status_var.set(
            f"已删除原顺序 {seq} 及关联记录{link_note}；后续顺序已前移"
        )
        return "break"

    def copy_title(self) -> None:
        selection = self.tree.selection()
        if selection:
            record = self.records.get(selection[0])
        else:
            children = self.tree.get_children()
            record = self.records.get(children[-1]) if children else None
        title = (record or {}).get("title")
        if not title:
            self.status_var.set("没有可复制的标题")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(title)
        self.status_var.set("标题已复制")

    def stop_task(self) -> None:
        """请求当前提取或更新任务安全停止。"""
        if not (self.running or self.refreshing):
            return
        self.cancel_event.set()
        self.stop_button.config(state="disabled")
        self.status_var.set("正在停止任务；已完成的记录会保留，当前暂存内容将清理…")

    def retry_failed(self) -> None:
        """只重新执行上一次失败的输入项。"""
        if self.running or self.refreshing:
            return
        jobs = list(self.failed_jobs)
        if not jobs:
            self.status_var.set("没有可重试的失败项")
            return
        self._retry_request = True
        self.start(jobs_override=jobs)

    def extract_current_once(self) -> None:
        """立即且只执行一次光标所在（或最近一条）链接的完整提取流程。"""
        if self.running or self.refreshing:
            self.status_var.set("正在提取/刷新中，请等当前任务完成后再单次提取")
            return
        text = self.input_text.get("1.0", "end")
        try:
            cursor_line = int(self.input_text.index("insert").split(".", 1)[0])
        except (AttributeError, TypeError, ValueError):
            cursor_line = len(text.splitlines()) or 1
        job = input_job_at_line(text, cursor_line)
        if job is None:
            self.status_var.set("当前及前面的条目中没有可单次提取的抖音链接")
            return
        self._enforce_input_sequences()
        self.start(jobs_override=[job])

    def open_covers(self) -> None:
        covers_dir = Path(self.output_var.get().strip() or default_output_dir()) / "封面"
        covers_dir.mkdir(parents=True, exist_ok=True)
        os.startfile(covers_dir)

    def open_log(self) -> None:
        log_path = Path(self.output_var.get().strip() or default_output_dir()) / LOG_NAME
        log_path.parent.mkdir(parents=True, exist_ok=True)
        if not log_path.exists():
            log_path.write_text("（暂无日志，点击「全部提取」后自动生成）\n", encoding="utf-8")
        os.startfile(log_path)

    def open_output(self) -> None:
        target = Path(self.output_var.get().strip() or default_output_dir())
        target.mkdir(parents=True, exist_ok=True)
        os.startfile(target)

    def open_records(self) -> None:
        """打开当前输出目录中已经存在的提取记录，不创建空工作簿。"""
        target = Path(self.output_var.get().strip() or default_output_dir()) / "提取记录.xlsx"
        if not target.is_file():
            self.status_var.set("输出目录里没有「提取记录.xlsx」")
            messagebox.showinfo(
                "没有提取记录",
                "当前输出目录里还没有「提取记录.xlsx」。\n请先完成一次提取。",
                parent=self.root,
            )
            return
        try:
            os.startfile(target)
            self.status_var.set("已打开「提取记录.xlsx」")
        except OSError as exc:
            self.status_var.set(f"无法打开提取记录：{exc}")
            messagebox.showerror(
                "无法打开提取记录",
                f"系统无法打开「提取记录.xlsx」：\n{exc}",
                parent=self.root,
            )

    def refresh_table_data(self) -> None:
        """按表格里的全部链接直接抓取并写入最新数据。"""
        if self.running or self.refreshing:
            self.status_var.set("正在提取/更新中，请等当前任务完成")
            return
        output_dir = Path(self.output_var.get().strip() or default_output_dir())
        xlsx = output_dir / "提取记录.xlsx"
        if not xlsx.exists():
            self.status_var.set("输出目录里没有「提取记录.xlsx」")
            return
        rows = exporter.read_records(xlsx)
        targets = [
            (seq, rec)
            for seq, rec in sorted(rows.items())
            if (rec.get("raw_input") or "").strip()
        ]
        if not targets:
            self.status_var.set("表格里没有带链接的行，无法刷新")
            return

        self._start_table_refresh(output_dir, targets)

    def _start_table_refresh(self, output_dir: Path, targets: list) -> None:
        """仅响应用户操作，启动安全的表格刷新。"""
        if self.refresh_thread is not None and self.refresh_thread.is_alive():
            self.status_var.set("上一次刷新仍在安全停止和清理浏览器，请稍候…")
            return
        setup_logger(output_dir / LOG_NAME)
        logging.getLogger("douyin_tool").info("手动刷新记录：共 %d 行有链接", len(targets))
        self.refreshing = True
        self.auto_refresh = False
        self.backup_enabled = bool(self.backup_var.get())
        self.cancel_event.clear()
        self.close_requested = False
        self.failed_jobs = []
        self.cancel_count = 0
        self.rollback_count = 0
        self._set_task_controls(False)
        self.update_button.config(text="刷新中…")
        self.status_var.set(f"开始刷新记录，共 {len(targets)} 行…")
        self.refresh_thread = threading.Thread(
            target=self._refresh_work_safe,
            args=(output_dir, targets),
            daemon=False,
        )
        self.refresh_thread.start()
        self.root.after(100, self._poll_refresh)

    def _refresh_work_safe(self, output_dir: Path, targets: list) -> None:
        """可取消、原子写回且补齐稳定 ID 的文档更新。"""
        logger = logging.getLogger("douyin_tool")
        updates: dict[int, dict] = {}
        failed_jobs: list[tuple[int | None, str]] = []
        paused_message = ""
        unchecked = 0
        terminal_kind = "rdone"
        terminal_payload = None
        access_context = extractor.AccessContext(
            BROWSER_PROFILE_DIR,
            self.cancel_event,
            lambda event, message: self._post(
                "rverification", {"event": event}, message
            ),
        )
        try:
            for index, (seq, rec) in enumerate(targets, 1):
                ensure_not_cancelled(self.cancel_event)
                self._post(
                    "rprogress", {"i": index, "n": len(targets), "seq": seq}
                )
                if index > 1:
                    interruptible_wait(1 + random.random(), self.cancel_event)
                link = rec.get("raw_input") or ""
                try:
                    fetched, fail_status = fetch_with_retry(
                        logger, seq, link, self.cancel_event, access_context
                    )
                except extractor.BrowserVerificationError as exc:
                    fetched, fail_status = None, exc.status
                    paused_message = f"批次已暂停：{exc}"
                    unchecked = len(targets) - index
                previous_status = (rec.get("status") or "").strip()
                success_status = "已恢复" if "失效" in previous_status else "正常"
                updates[seq] = {
                    "record": fetched,
                    "status": success_status if fetched is not None else fail_status,
                    "link": link,
                }
                # 每检查完一条就立即写回，避免整批结束前 Excel 仍显示旧数据。
                rows = exporter.read_records(output_dir / "提取记录.xlsx")
                current = dict(rows.get(seq) or rec)
                if fetched is not None:
                    current.update(
                        {
                            "raw_input": link,
                            "title": fetched.fields["title"],
                            "tags": fetched.fields["tags"],
                            "likes": fetched.fields["likes"],
                            "comments": fetched.fields["comments"],
                            "author": fetched.fields["author"],
                            "status": success_status,
                            "aweme_id": fetched.aweme_id,
                            "work_kind": "图文" if fetched.kind == "note" else "视频",
                            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                    )
                else:
                    logger.warning(
                        "顺序 %d 更新失败（%s），保留旧数据并更新状态",
                        seq,
                        fail_status,
                    )
                    current["status"] = fail_status
                    current["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                rows[seq] = current
                update_records_force_close(
                    output_dir / "提取记录.xlsx",
                    rows,
                    [seq],
                    allow_force_close=not self.auto_refresh,
                    keep_backup=getattr(self, "backup_enabled", False),
                )
                self._post(
                    "rresult",
                    {"seq": seq, "success": fetched is not None, "record": current},
                    None if fetched is not None else fail_status,
                )
                if fetched is None:
                    failed_jobs.append((seq, link))
                else:
                    logger.info(
                        "顺序 %d 更新成功，作品 ID %s，点赞 %d，评论 %d",
                        seq,
                        fetched.aweme_id,
                        fetched.fields["likes"],
                        fetched.fields["comments"],
                    )
                if paused_message:
                    logger.warning(
                        "浏览器验证未完成，刷新熔断：当前失败 1 条，未检查 %d 条",
                        unchecked,
                    )
                    break

            ok_count = sum(1 for update in updates.values() if update["record"] is not None)
            terminal_payload = {
                "ok": ok_count,
                "fail": len(updates) - ok_count,
                "total": len(updates),
                "failed_jobs": failed_jobs,
                "unchecked": unchecked,
                "message": paused_message,
            }
        except TaskCancelled:
            logger.info("刷新记录已按用户请求取消")
            terminal_kind = "rcancelled"
            terminal_payload = {
                "processed": len(updates),
                "total": len(targets),
                "failed_jobs": failed_jobs,
            }
        except Exception as exc:
            logger.exception("刷新记录异常终止")
            terminal_kind = "rerror"
            terminal_payload = str(exc)
        finally:
            # 终态消息必须晚于浏览器/Playwright 清理。否则主线程会提前解锁
            # “刷新记录”，第二个任务可能与仍在退出的持久化浏览器争用资料目录。
            access_context.close()
            self._post(terminal_kind, terminal_payload)

    def _poll_refresh(self) -> None:
        try:
            while True:
                kind, payload, extra = self.message_queue.get_nowait()
                if kind == "rprogress":
                    self.progress["value"] = int(payload["i"] / max(1, payload["n"]) * 100)
                    self.progress_label.config(
                        text=f"刷新 {payload['i']}/{payload['n']}（顺序 {payload['seq']}）"
                    )
                    self._highlight_record(payload["seq"], "refreshing")
                elif kind == "rresult":
                    tag = "refresh_success" if payload["success"] else "refresh_failure"
                    self._update_visible_record(payload["seq"], payload["record"], tag)
                    if not payload["success"]:
                        self.status_var.set(
                            f"顺序 {payload['seq']} 更新失败：{extra or '获取失败'}"
                        )
                elif kind == "rverification":
                    self.status_var.set(extra or "请在弹出的浏览器中完成验证")
                    event = (payload or {}).get("event") if isinstance(payload, dict) else ""
                    self.progress_label.config(
                        text=(
                            "网络链路切换中…"
                            if event in {"network_fallback", "network_retry"}
                            else "等待浏览器验证…"
                        )
                    )
                elif kind == "rdone":
                    self.refreshing = False
                    self.auto_refresh = False
                    self.failed_jobs = list(payload.get("failed_jobs") or [])
                    self._set_task_controls(True)
                    self.update_button.config(text="刷新记录")
                    self.progress["value"] = 0
                    self.progress_label.config(text="")
                    self.load_existing_records()
                    self.status_var.set(
                        payload.get("message")
                        or (
                            "刷新完成"
                            + f"：共 {payload['total']} 行"
                            f"（正常 {payload['ok']}，异常 {payload['fail']}）"
                        )
                    )
                    if payload.get("message"):
                        self.status_var.set(
                            f"{payload['message']}；已检查 {payload['total']}，"
                            f"正常 {payload['ok']}，异常 {payload['fail']}，"
                            f"未检查 {payload.get('unchecked', 0)}"
                        )
                    if self.close_requested:
                        self.root.after(50, self._wait_then_close)
                    return
                elif kind == "rcancelled":
                    self.refreshing = False
                    self.auto_refresh = False
                    self.failed_jobs = list(payload.get("failed_jobs") or [])
                    self._set_task_controls(True)
                    self.update_button.config(text="刷新记录")
                    self.progress["value"] = 0
                    self.progress_label.config(text="")
                    self.status_var.set(
                        f"更新已取消：已检查 {payload.get('processed', 0)}/{payload.get('total', 0)} 行"
                    )
                    if self.close_requested:
                        self.root.after(50, self._wait_then_close)
                    return
                elif kind == "rerror":
                    self.refreshing = False
                    self.auto_refresh = False
                    self._set_task_controls(True)
                    self.update_button.config(text="刷新记录")
                    self.progress["value"] = 0
                    self.progress_label.config(text="")
                    self.status_var.set(f"刷新失败：{payload}")
                    if self.close_requested:
                        self.root.after(50, self._wait_then_close)
                    return
        except queue.Empty:
            pass
        if self.refreshing:
            self.root.after(100, self._poll_refresh)

    def _schedule_renumber(self, _event=None) -> None:
        """输入变化后统一调度：锁定序号并维护末尾待填编号。"""
        if self._renumber_after_id is not None:
            self.root.after_cancel(self._renumber_after_id)
        self._renumber_after_id = self.root.after_idle(self._renumber_input)

    def _on_input_paste(self, _event=None):
        """粘贴发生前立即检查上方输入和下方已提取记录。"""
        try:
            pasted = self.root.clipboard_get()
        except tk.TclError:
            self._schedule_renumber()
            return None

        current = self.input_text.get("1.0", "end-1c")
        try:
            before = self.input_text.get("1.0", "sel.first")
            after = self.input_text.get("sel.last", "end-1c")
            current = before + after
        except tk.TclError:
            pass

        duplicates = input_parser.existing_duplicate_urls(current, str(pasted))
        known = {input_parser.link_identity(url) for url in extractor.extract_urls(current)}
        lower_duplicates: list[tuple[str, int]] = []
        for url in extractor.extract_urls(str(pasted)):
            identity = input_parser.link_identity(url)
            if identity in known:
                continue
            for record in getattr(self, "records", {}).values():
                record_urls = extractor.extract_urls(str(record.get("raw_input") or ""))
                if any(input_parser.link_identity(value) == identity for value in record_urls):
                    lower_duplicates.append((url, int(record.get("seq") or 0)))
                    known.add(identity)
                    break
        duplicates.extend(lower_duplicates)
        if duplicates:
            details = "\n".join(
                f"· 第 {seq} 条：{url}" for url, seq in duplicates[:5]
            )
            if len(duplicates) > 5:
                details += f"\n· 其他 {len(duplicates) - 5} 条重复链接"
            self.status_var.set(
                f"链接重复：已在第 {duplicates[0][1]} 条，本次未粘贴"
            )
            messagebox.showwarning(
                "链接重复",
                f"本次粘贴的链接已在列表中，已阻止重复添加：\n\n{details}",
                parent=self.root,
            )
            return "break"

        affected = self._selected_input_jobs()
        if any(self._matching_record_for_input_job(seq, raw) for seq, raw in affected):
            self.status_var.set("选区包含已提取链接，请先删除该链接再粘贴")
            messagebox.showwarning(
                "不能直接替换已提取链接",
                "当前选区包含已提取链接。\n\n"
                "请先使用“删除当前链接”同步删除对应记录，再粘贴新链接。",
                parent=self.root,
            )
            return "break"

        self._schedule_renumber()
        return None

    def _matching_record_for_input_job(self, seq: int, raw: str):
        """按输入出现位置定位记录；重复链接绝不回退到前面的第一条。"""
        identities = {
            input_parser.link_identity(url) for url in extractor.extract_urls(raw)
        }
        if not identities:
            return None
        exact = []
        candidates = []
        for item_id, record in self.records.items():
            record_ids = {
                input_parser.link_identity(url)
                for url in extractor.extract_urls(str(record.get("raw_input") or ""))
            }
            if not identities.intersection(record_ids):
                continue
            candidates.append((item_id, record))
            try:
                record_seq = int(record.get("seq"))
            except (TypeError, ValueError):
                record_seq = 0
            if record_seq == seq:
                exact.append((item_id, record))
        if len(exact) == 1:
            return exact[0]
        if len(candidates) == 1:
            return candidates[0]
        return None

    def _selected_input_jobs(self) -> list[tuple[int, str]]:
        """返回当前选区实际覆盖到 URL 的输入条目。"""
        try:
            start = self.input_text.index("sel.first")
            end = self.input_text.index("sel.last")
        except tk.TclError:
            return []
        start_line, start_col = (int(value) for value in start.split(".", 1))
        end_line, end_col = (int(value) for value in end.split(".", 1))
        current = self.input_text.get("1.0", "end-1c")
        found: list[tuple[int, str]] = []
        seen: set[int] = set()
        for line_number in range(start_line, end_line + 1):
            line = self.input_text.get(f"{line_number}.0", f"{line_number}.end")
            selection_start = start_col if line_number == start_line else 0
            selection_end = end_col if line_number == end_line else len(line)
            if not any(
                match.start() < selection_end and match.end() > selection_start
                for match in extractor.DOUYIN_URL_RE.finditer(line)
            ):
                continue
            job = input_parser.input_job_at_line(current, line_number)
            if job and job[0] not in seen:
                found.append(job)
                seen.add(job[0])
        return found

    def _input_jobs_affected_by_delete_key(self, event) -> list[tuple[int, str]]:
        selected = self._selected_input_jobs()
        if selected:
            return selected
        try:
            index = self.input_text.index("insert")
            line_number, column = (int(value) for value in index.split(".", 1))
            line = self.input_text.get(f"{line_number}.0", f"{line_number}.end")
        except (TypeError, ValueError, tk.TclError):
            return []
        character = column - 1 if getattr(event, "keysym", "") == "BackSpace" else column
        if character < 0 or not any(
            match.start() <= character < match.end()
            for match in extractor.DOUYIN_URL_RE.finditer(line)
        ):
            return []
        current = self.input_text.get("1.0", "end-1c")
        job = input_parser.input_job_at_line(current, line_number)
        return [job] if job else []

    def _route_input_jobs_to_record_deletion(self, jobs: list[tuple[int, str]]) -> bool:
        """在文本被改坏前，将单个已提取条目送入完整删除事务。"""
        matches = []
        for seq, raw in jobs:
            match = self._matching_record_for_input_job(seq, raw)
            if match and match not in matches:
                matches.append(match)
        if not matches:
            return False
        if len(matches) != 1:
            self.status_var.set("一次只能同步删除一条已提取链接")
            messagebox.showwarning(
                "请逐条删除",
                "选区包含多条已提取链接。为避免误删，请每次只删除一条。",
                parent=self.root,
            )
            return True
        item_id, record = matches[0]
        self.tree.selection_set(item_id)
        self.tree.focus(item_id)
        self.tree.see(item_id)
        self.status_var.set(f"请确认同步删除第 {record.get('seq')} 条已提取链接")
        self.delete_selected_record()
        return True

    def _on_input_delete_key(self, event):
        if self.running or self.refreshing:
            return None
        jobs = self._input_jobs_affected_by_delete_key(event)
        return "break" if self._route_input_jobs_to_record_deletion(jobs) else None

    def _on_input_cut(self, _event=None):
        if self.running or self.refreshing:
            return None
        if self._route_input_jobs_to_record_deletion(self._selected_input_jobs()):
            return "break"
        self._schedule_renumber()
        return None

    def _renumber_input(self) -> None:
        """锁定序号 + 追加下一个待填编号。

        - 序号永远等于条目位置：把 5. 改成 9. 也会恢复成 5.；
        - 只规范/恢复序号前缀，后面的原始链接与分享文案原样保留；
        - 多行粘贴仍视为一条内容，不逐行重排。
        """
        self._enforce_input_sequences()

    def clear_input(self) -> None:
        self.input_text.delete("1.0", "end")
        self.input_text.insert("1.0", "1.")
        self._style_input_sequences()
        self._save_input_cache()
        self.status_var.set("输入已清空")

    def delete_current_input_link(self, _event=None) -> str:
        """删除光标所在的上方输入链接，只将后续序号前移。"""
        if self.running or self.refreshing:
            self.status_var.set("正在提取/刷新中，不能删除输入链接")
            return "break"
        try:
            line_number = int(self.input_text.index("insert").split(".", 1)[0])
        except (TypeError, ValueError, tk.TclError):
            line_number = 1
        current = self.input_text.get("1.0", "end-1c")
        updated, removed_seq, removed_raw = input_parser.remove_entry_at_line(
            current, line_number
        )
        if removed_seq is None:
            self.status_var.set("请先把光标放在要删除的链接内容上")
            return "break"

        match = self._matching_record_for_input_job(removed_seq, removed_raw)
        if match:
            item_id, _record = match
            self.tree.selection_set(item_id)
            self.tree.focus(item_id)
            self.tree.see(item_id)
            return self.delete_selected_record()

        self.input_text.delete("1.0", "end")
        self.input_text.insert("1.0", updated)
        self.input_text.mark_set("insert", "end-1c")
        self.input_text.see("insert")
        self._style_input_sequences()
        self._save_input_cache()
        self.status_var.set(
            f"已删除第 {removed_seq} 条输入链接，后续序号已依次前移"
        )
        return "break"


def report_unhandled_error(
    exc_type: type[BaseException],
    exc_value: BaseException,
    exc_traceback,
    parent: tk.Misc | None = None,
) -> None:
    """无控制台运行时把未处理异常写入文件，并通过图形弹窗告知用户。"""
    details = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    try:
        STARTUP_ERROR_FILE.parent.mkdir(parents=True, exist_ok=True)
        with STARTUP_ERROR_FILE.open("a", encoding="utf-8") as handle:
            handle.write(f"\n[{time.strftime('%Y-%m-%d %H:%M:%S')}]\n{details}")
    except Exception:
        pass
    try:
        messagebox.showerror(
            "抖音信息提取工具",
            f"程序遇到错误：{exc_value}\n\n详细信息已写入：\n{STARTUP_ERROR_FILE}",
            parent=parent,
        )
    except Exception:
        pass


def main() -> None:
    root: tk.Tk | None = None
    try:
        root = tk.Tk()
        root.report_callback_exception = lambda exc_type, exc_value, exc_traceback: (
            report_unhandled_error(exc_type, exc_value, exc_traceback, root)
        )
        DouyinExtractorApp(root)
        root.mainloop()
    except Exception:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        if exc_type is not None and exc_value is not None:
            report_unhandled_error(exc_type, exc_value, exc_traceback, root)


if __name__ == "__main__":
    main()
