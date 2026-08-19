# -*- coding: utf-8 -*-
"""安全更新提取记录工作簿，并保留用户的其它工作表和自定义内容。"""

from __future__ import annotations

from datetime import datetime
import io
import os
from pathlib import Path
import shutil
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

BASE_HEADERS = [
    "视频链接",
    "标题",
    "标签",
    "点赞数",
    "评论数",
    "类型",
    "博主",
    "状态",
    "封面",
    "顺序",
]
HEADERS = BASE_HEADERS + ["作品ID", "作品形式", "最后更新"]
OLD_HEADERS = ["视频链接", "标题", "标签", "点赞数", "评论数", "类型", "博主", "封面", "顺序"]
COLUMN_WIDTHS = [58, 46, 30, 10, 10, 14, 22, 18, 12, 8, 24, 12, 20]
IMAGE_HEIGHT_PX = 84
SHEET_NAME = "提取记录"
_KEEP_COVER = object()


class WorkbookInUseError(PermissionError):
    """目标工作簿被 Excel/WPS 占用。"""


def _cover_png_bytes(cover_path) -> tuple[io.BytesIO, int, int]:
    """把封面转成 PNG 字节流。"""
    with PILImage.open(cover_path) as source:
        image = source.copy()
    image.thumbnail((480, 480))
    if image.mode != "RGB":
        image = image.convert("RGB")
    width, height = image.size
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer, width, height


def _headers(sheet) -> list[str]:
    if sheet.max_row < 1:
        return []
    return [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]


def _looks_managed(sheet) -> bool:
    headers = set(_headers(sheet))
    return "视频链接" in headers and "顺序" in headers


def _select_sheet(workbook, create: bool):
    if SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SHEET_NAME]
        if _looks_managed(sheet) or sheet.max_row <= 1:
            return sheet
    for sheet in workbook.worksheets:
        if _looks_managed(sheet):
            return sheet
    if not create:
        return None
    if len(workbook.worksheets) == 1:
        sheet = workbook.active
        if sheet.max_row <= 1 and not any(_headers(sheet)):
            sheet.title = SHEET_NAME
            return sheet
    title = SHEET_NAME
    if title in workbook.sheetnames:
        suffix = 2
        while f"{title}{suffix}" in workbook.sheetnames:
            suffix += 1
        title = f"{title}{suffix}"
    return workbook.create_sheet(title)


def _ensure_headers(sheet) -> dict[str, int]:
    current = _headers(sheet)
    if not any(current):
        current = []
    for name in HEADERS:
        if name not in current:
            current.append(name)
    for column, text in enumerate(current, 1):
        cell = sheet.cell(row=1, column=column, value=text)
        if text in HEADERS:
            cell.font = Font(bold=True)
            width = COLUMN_WIDTHS[HEADERS.index(text)]
            sheet.column_dimensions[get_column_letter(column)].width = width
    return {name: index + 1 for index, name in enumerate(current) if name}


def _parse_seq(value) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and float(value).is_integer():
        seq = int(value)
        return seq if seq > 0 else None
    if isinstance(value, str) and value.strip().isdigit():
        seq = int(value.strip())
        return seq if seq > 0 else None
    return None


def _record_from_row(values: list, layout: dict[str, int]) -> tuple[int | None, dict]:
    def pick(name: str):
        column = layout.get(name)
        index = column - 1 if column else None
        return values[index] if index is not None and index < len(values) else None

    seq = _parse_seq(pick("顺序"))
    record = {
        "raw_input": pick("视频链接") or "",
        "title": pick("标题") or "",
        "tags": pick("标签") or "",
        "likes": int(pick("点赞数") or 0),
        "comments": int(pick("评论数") or 0),
        "type": pick("类型") or "基本盘",
        "author": pick("博主") or "",
        "status": pick("状态") or "",
        "aweme_id": str(pick("作品ID") or "").strip(),
        "work_kind": pick("作品形式") or "",
        "updated_at": pick("最后更新") or "",
    }
    return seq, record


def read_records(path) -> dict[int, dict]:
    """按表头名称读取新旧工作簿，未知格式不做破坏性猜测。"""
    path = Path(path)
    if not path.exists():
        return {}
    workbook = load_workbook(path)
    try:
        sheet = _select_sheet(workbook, create=False)
        if sheet is None:
            return {}
        layout = {name: index + 1 for index, name in enumerate(_headers(sheet)) if name}
        result: dict[int, dict] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            seq, record = _record_from_row(list(row), layout)
            if seq is not None:
                result[seq] = record
        return result
    finally:
        workbook.close()


def _image_row(image) -> int | None:
    try:
        return int(image.anchor._from.row) + 1
    except (AttributeError, TypeError, ValueError):
        return None


def _replace_cover(sheet, row: int, cover_path, cover_column: int) -> None:
    if cover_path is _KEEP_COVER:
        return
    sheet._images = [image for image in sheet._images if _image_row(image) != row]
    sheet.cell(row=row, column=cover_column, value=None)
    if not cover_path or not Path(cover_path).exists():
        return
    buffer, width, height = _cover_png_bytes(cover_path)
    image = XLImage(buffer)
    image.height = IMAGE_HEIGHT_PX
    image.width = max(1, round(width / max(1, height) * IMAGE_HEIGHT_PX))
    image.anchor = f"{get_column_letter(cover_column)}{row}"
    sheet.add_image(image)
    sheet.row_dimensions[row].height = 68


def _write_record(sheet, row: int, record: dict, seq: int, layout: dict[str, int], cover_path) -> None:
    values = {
        "视频链接": record.get("raw_input") or "",
        "标题": record.get("title") or "",
        "标签": record.get("tags") or "",
        "点赞数": int(record.get("likes") or 0),
        "评论数": int(record.get("comments") or 0),
        "类型": record.get("type") or "基本盘",
        "博主": record.get("author") or "",
        "状态": record.get("status") or "",
        "顺序": seq,
        "作品ID": str(record.get("aweme_id") or ""),
        "作品形式": record.get("work_kind") or "",
        "最后更新": record.get("updated_at") or "",
    }
    for name, value in values.items():
        column = layout[name]
        cell = sheet.cell(row=row, column=column, value=value)
        cell.alignment = Alignment(vertical="center", wrap_text=name in {"视频链接", "标题", "标签"})
    _replace_cover(sheet, row, cover_path, layout["封面"])


def _backup_workbook(path: Path) -> None:
    if not path.exists():
        return
    backup_dir = path.parent / "表格备份"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S_%f")
    shutil.copy2(path, backup_dir / f"{path.stem}_{timestamp}{path.suffix}")
    backups = sorted(backup_dir.glob(f"{path.stem}_*{path.suffix}"), key=lambda p: p.stat().st_mtime)
    for old in backups[:-5]:
        old.unlink(missing_ok=True)


def _save_atomic(workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp{path.suffix}")
    try:
        workbook.save(temp)
        check = load_workbook(temp, read_only=True)
        check.close()
        _backup_workbook(path)
        os.replace(temp, path)
    except PermissionError as exc:
        raise WorkbookInUseError(
            "提取记录.xlsx 正被 Excel/WPS 占用，请关闭该工作簿后重试"
        ) from exc
    finally:
        temp.unlink(missing_ok=True)


def update_records(
    path,
    records: dict[int, dict],
    seqs: list[int] | None = None,
    cover_map: dict[int, str] | None = None,
) -> None:
    """只更新工具管理的行和列，保留其它工作表、额外列和格式。"""
    path = Path(path)
    workbook = load_workbook(path) if path.exists() else Workbook()
    try:
        sheet = _select_sheet(workbook, create=True)
        layout = _ensure_headers(sheet)
        row_by_seq: dict[int, int] = {}
        seq_column = layout["顺序"]
        for row in range(2, sheet.max_row + 1):
            seq = _parse_seq(sheet.cell(row=row, column=seq_column).value)
            if seq is not None:
                row_by_seq[seq] = row

        selected = seqs if seqs is not None else sorted(records)
        covers = cover_map or {}
        for seq in selected:
            record = records.get(seq)
            if record is None:
                continue
            row = row_by_seq.get(seq)
            if row is None:
                row = sheet.max_row + 1
                row_by_seq[seq] = row
            cover_value = covers[seq] if seq in covers else _KEEP_COVER
            _write_record(sheet, row, record, seq, layout, cover_value)
        _save_atomic(workbook, path)
    finally:
        workbook.close()


def append_record(path, record: dict, cover_path, seq: int) -> int:
    """兼容旧调用：安全新增或更新一条记录。"""
    update_records(path, {seq: record}, [seq], {seq: str(cover_path)} if cover_path else {})
    rows = read_records(path)
    return sorted(rows).index(seq) + 2 if seq in rows else 0


def rebuild_workbook(
    path, records: dict[int, dict], seqs: list[int], cover_map: dict[int, str]
) -> None:
    """兼容旧名称：执行非破坏性的原子更新，不再重建整个工作簿。"""
    update_records(path, records, seqs, cover_map)
