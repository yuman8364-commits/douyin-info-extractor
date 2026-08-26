# -*- coding: utf-8 -*-

from pathlib import Path
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook
from PIL import Image

import exporter


class ExporterTests(unittest.TestCase):
    def _legacy_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "提取记录"
        sheet.append(exporter.BASE_HEADERS + ["人工备注"])
        sheet.append(
            ["https://v.douyin.com/Old/", "旧标题", "#旧", 1, 2, "人工分类", "作者", "正常", None, 1, "别覆盖"]
        )
        other = workbook.create_sheet("我的分析")
        other["A1"] = "保留这个工作表"
        workbook.save(path)
        workbook.close()

    def test_migration_preserves_extra_sheet_column_and_manual_type(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "提取记录.xlsx"
            self._legacy_workbook(path)
            records = exporter.read_records(path)
            records[1].update(
                {
                    "title": "新标题",
                    "likes": 99,
                    "aweme_id": "123",
                    "work_kind": "视频",
                    "updated_at": "2026-08-19 15:00:00",
                }
            )
            exporter.update_records(path, records, [1])

            workbook = load_workbook(path)
            self.assertIn("我的分析", workbook.sheetnames)
            self.assertEqual(workbook["我的分析"]["A1"].value, "保留这个工作表")
            sheet = workbook["提取记录"]
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("作品ID", headers)
            self.assertIn("作品形式", headers)
            self.assertIn("最后更新", headers)
            self.assertEqual(sheet.cell(2, headers.index("类型") + 1).value, "人工分类")
            self.assertEqual(sheet.cell(2, headers.index("人工备注") + 1).value, "别覆盖")
            workbook.close()
            self.assertTrue(any((path.parent / "表格备份").glob("*.xlsx")))

    def test_replace_failure_keeps_original_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "提取记录.xlsx"
            self._legacy_workbook(path)
            original = path.read_bytes()
            records = exporter.read_records(path)
            records[1]["title"] = "不应落盘"
            with mock.patch.object(exporter.os, "replace", side_effect=PermissionError("locked")):
                with self.assertRaises(exporter.WorkbookInUseError):
                    exporter.update_records(path, records, [1])
            self.assertEqual(path.read_bytes(), original)
            self.assertEqual(list(path.parent.glob(".*.tmp.xlsx")), [])

    def test_workbook_backup_can_be_disabled(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "提取记录.xlsx"
            self._legacy_workbook(path)
            records = exporter.read_records(path)
            records[1]["title"] = "不保留表格备份"

            exporter.update_records(path, records, [1], keep_backup=False)

            self.assertFalse((path.parent / "表格备份").exists())
            self.assertEqual(exporter.read_records(path)[1]["title"], "不保留表格备份")

    def test_delete_record_removes_row_and_shifts_following_sequences(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "提取记录.xlsx"
            records = {
                1: {"raw_input": "https://example.com/1", "title": "第一条"},
                2: {"raw_input": "https://example.com/2", "title": "删除我"},
                3: {"raw_input": "https://example.com/3", "title": "原第三条"},
            }
            exporter.update_records(path, records, keep_backup=False)

            deleted = exporter.delete_record(path, 2, keep_backup=False)

            self.assertTrue(deleted)
            rows = exporter.read_records(path)
            self.assertEqual(sorted(rows), [1, 2])
            self.assertEqual(rows[2]["title"], "原第三条")
            self.assertFalse((path.parent / "表格备份").exists())

    def test_delete_record_removes_cover_and_shifts_later_image_anchor(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / "提取记录.xlsx"
            covers = {}
            for seq in (1, 2, 3):
                cover = root / f"{seq}.jpg"
                Image.new("RGB", (20, 30), (seq * 40, 0, 0)).save(cover)
                covers[seq] = str(cover)
            records = {seq: {"title": f"第 {seq} 条"} for seq in (1, 2, 3)}
            exporter.update_records(path, records, cover_map=covers, keep_backup=False)

            self.assertTrue(exporter.delete_record(path, 2, keep_backup=False))

            workbook = load_workbook(path)
            sheet = workbook["提取记录"]
            image_rows = sorted(exporter._image_row(image) for image in sheet._images)
            self.assertEqual(image_rows, [2, 3])
            workbook.close()

    def test_delete_missing_record_keeps_workbook_unchanged(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "提取记录.xlsx"
            exporter.update_records(path, {1: {"title": "保留"}}, keep_backup=False)
            original = path.read_bytes()

            self.assertFalse(exporter.delete_record(path, 9, keep_backup=False))
            self.assertEqual(path.read_bytes(), original)


if __name__ == "__main__":
    unittest.main()
